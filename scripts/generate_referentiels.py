#!/usr/bin/env python3
"""
Générateur du catalogue Référentiels CyberRepères.

Source par défaut :
    catalogues/CyberReperes_CatalogueRef_Cyber.xlsx

Sorties par défaut :
    referentiels/index.md
    downloads/CyberReperes_Catalogue_Referentiels.xlsx

Commandes :
    python scripts/generate_referentiels.py --check
    python scripts/generate_referentiels.py
"""

from __future__ import annotations

import argparse
import html
import shutil
import sys
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


DEFAULT_EXCEL = Path("catalogues/CyberReperes_CatalogueRef_Cyber.xlsx")
DEFAULT_OUTPUT = Path("referentiels/index.md")
DEFAULT_DOWNLOAD = Path("downloads/CyberReperes_Catalogue_Referentiels.xlsx")

REQUIRED_COLUMNS = [
    "ID",
    "Organisme",
    "Titre",
    "Type",
    "Domaine",
    "Périmètre",
    "Langue",
    "Année",
    "Version",
    "Statut",
    "Description succincte",
    "Source officielle",
    "Téléchargement",
    "Accès",
    "Pertinence",
    "Notions",
    "Méthodes",
    "Livrables",
    "Référentiels liés",
    "Mots-clés",
    "Notes CyberRepères",
]


def clean(value) -> str:
    """Convertit une cellule Excel en texte propre."""
    if value is None:
        return ""
    return str(value).strip()


def is_url(value: str) -> bool:
    """Retourne True si la valeur est une URL HTTP(S) valide."""
    if not value:
        return False
    try:
        parsed = urlparse(value)
        return parsed.scheme in ("http", "https") and bool(parsed.netloc)
    except Exception:
        return False


def md_escape(value: str) -> str:
    """Échappe les caractères problématiques dans les cellules Markdown."""
    return clean(value).replace("|", r"\|").replace("\n", " ")


def html_text(value: str) -> str:
    """Nettoie et échappe une valeur destinée au HTML."""
    value = clean(value)
    value = " ".join(value.split())
    return html.escape(value, quote=True)


def html_link(label: str, url: str) -> str:
    """Construit un lien HTML sûr."""
    if not is_url(url):
        return ""
    return (
        f'<a href="{html_text(url)}" target="_blank" rel="noopener">'
        f"{html_text(label)}</a>"
    )


def load_catalogue(path: Path) -> list[dict[str, str]]:
    """Lit l'onglet Catalogue du fichier Excel."""
    if not path.exists():
        raise FileNotFoundError(f"Fichier Excel introuvable : {path}")

    workbook = load_workbook(path, data_only=True)

    if "Catalogue" not in workbook.sheetnames:
        raise ValueError("L'onglet 'Catalogue' est absent du fichier Excel.")

    sheet = workbook["Catalogue"]

    headers = [
        clean(sheet.cell(1, column).value)
        for column in range(1, sheet.max_column + 1)
    ]

    positions = {
        header: index + 1
        for index, header in enumerate(headers)
        if header
    }

    missing = [column for column in REQUIRED_COLUMNS if column not in positions]

    if missing:
        raise ValueError(
            "Colonnes obligatoires absentes : " + ", ".join(missing)
        )

    rows = []

    for row_number in range(2, sheet.max_row + 1):
        row = {
            column: clean(sheet.cell(row_number, positions[column]).value)
            for column in REQUIRED_COLUMNS
        }

        # Ignore les lignes totalement vides.
        if not row["ID"] and not row["Titre"]:
            continue

        rows.append(row)

    return rows


def validate(rows: list[dict[str, str]]) -> tuple[list[str], list[str]]:
    """Contrôle le catalogue avant génération."""
    errors = []
    warnings = []
    seen_ids = set()

    for row in rows:
        identifier = row["ID"] or row["Titre"] or "(sans identifiant)"

        if not row["ID"]:
            errors.append(f"ID manquant : {identifier}")
        elif row["ID"] in seen_ids:
            errors.append(f"ID en doublon : {row['ID']}")

        seen_ids.add(row["ID"])

        for field in (
            "Organisme",
            "Titre",
            "Type",
            "Description succincte",
        ):
            if not row[field]:
                errors.append(f"{field} manquant : {identifier}")

        if not row["Source officielle"]:
            warnings.append(f"URL officielle manquante : {identifier}")
        elif not is_url(row["Source officielle"]):
            errors.append(
                f"URL officielle invalide : {identifier} "
                f"({row['Source officielle']})"
            )

        if row["Téléchargement"] and not is_url(row["Téléchargement"]):
            warnings.append(
                f"Téléchargement non-URL à vérifier : {identifier} "
                f"({row['Téléchargement']})"
            )

    return errors, warnings


def detail_block(row: dict[str, str]) -> str:
    """Construit le contenu affiché lorsque la ligne est dépliée."""
    parts = []

    if row["Description succincte"]:
        parts.append(
            "<p><strong>Description :</strong> "
            f"{html_text(row['Description succincte'])}</p>"
        )

    fields = [
        ("Langue", "Langue"),
        ("Version", "Version"),
        ("Statut", "Statut"),
        ("Accès", "Accès"),
        ("Pertinence", "Pertinence"),
        ("Notions", "Notions"),
        ("Méthodes", "Méthodes"),
        ("Livrables", "Livrables"),
        ("Référentiels liés", "Référentiels liés"),
        ("Mots-clés", "Mots-clés"),
    ]

    items = []

    for label, key in fields:
        if row[key]:
            items.append(
                f"<li><strong>{html_text(label)} :</strong> "
                f"{html_text(row[key])}</li>"
            )

    if items:
        parts.append("<ul>" + "".join(items) + "</ul>")

    links = []

    if row["Source officielle"]:
        links.append(html_link("Source officielle", row["Source officielle"]))

    if (
        row["Téléchargement"]
        and is_url(row["Téléchargement"])
        and row["Téléchargement"] != row["Source officielle"]
    ):
        links.append(html_link("Télécharger", row["Téléchargement"]))

    links = [item for item in links if item]

    if links:
        parts.append("<p>" + " · ".join(links) + "</p>")

    if row["Notes CyberRepères"]:
        parts.append(
            "<p><strong>Note CyberRepères :</strong> "
            f"{html_text(row['Notes CyberRepères'])}</p>"
        )

    return "\n".join(parts)


def generate(
    rows: list[dict[str, str]],
    output: Path,
    download: Path,
    source_excel: Path,
) -> None:
    """Génère index.md et la copie publique du catalogue Excel."""

    rows = sorted(
        rows,
        key=lambda row: (
            row["Organisme"].lower(),
            row["Domaine"].lower(),
            row["Titre"].lower(),
        ),
    )

    organisms = sorted(
        {row["Organisme"] for row in rows if row["Organisme"]},
        key=str.lower,
    )

    lines = [
        "---",
        "layout: default",
        'title: "Référentiels"',
        'description: "Catalogue des guides, publications, recommandations et autres documents de référence sélectionnés par CyberRepères."',
        "nav: true",
        "nav_order: 20",
        "---",
        "",
        "# Référentiels",
        "",
        "Cette sélection rassemble des guides, publications, recommandations et autres documents de référence utiles à la cybersécurité industrielle et à ses domaines connexes.",
        "",
        f"**{len(rows)} références** · **{len(organisms)} organismes**",
        "",
        f"📥 [Télécharger le catalogue Excel]({{{{ site.baseurl }}}}/{download.as_posix()})",
        "",
        '<div class="catalogue-table">',
        '  <div class="catalogue-header">',
        "    <span>Référence</span>",
        "    <span>Document</span>",
        "    <span>Organisme</span>",
        "    <span>Type</span>",
        "    <span>Périmètre</span>",
        "    <span>Année</span>",
        "    <span>Source</span>",
        "  </div>",
    ]

    for row in rows:
        year = row["Année"] or "—"
        source = (
            html_link("🔗", row["Source officielle"])
            if row["Source officielle"]
            else "—"
        )

        detail = detail_block(row)

        lines.extend(
            [
                '  <details class="catalogue-entry">',
                '    <summary class="catalogue-row">',
                f'      <span class="catalogue-ref"><strong>{html_text(row["ID"])}</strong></span>',
                f'      <span>{html_text(row["Titre"])}</span>',
                f'      <span>{html_text(row["Organisme"])}</span>',
                f'      <span>{html_text(row["Type"])}</span>',
                f'      <span>{html_text(row["Périmètre"])}</span>',
                f'      <span>{html_text(year)}</span>',
                f"      <span>{source}</span>",
                "    </summary>",
                f'    <div class="catalogue-detail">{detail}</div>',
                "  </details>",
            ]
        )

    lines.extend(
        [
            "</div>",
            "",
        ]
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text("\n".join(lines), encoding="utf-8")

    download.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_excel, download)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Génère le catalogue Référentiels CyberRepères."
    )

    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_EXCEL,
        help="Fichier Excel source.",
    )

    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Fichier Markdown généré.",
    )

    parser.add_argument(
        "--download",
        type=Path,
        default=DEFAULT_DOWNLOAD,
        help="Copie publique du catalogue Excel.",
    )

    parser.add_argument(
        "--check",
        action="store_true",
        help="Vérifie le catalogue sans générer les fichiers.",
    )

    args = parser.parse_args()

    try:
        rows = load_catalogue(args.excel)
        errors, warnings = validate(rows)
    except Exception as exc:
        print(f"ERREUR : {exc}", file=sys.stderr)
        return 1

    print(f"Références lues : {len(rows)}")

    for warning in warnings:
        print(f"ATTENTION : {warning}")

    if errors:
        print("\nErreurs bloquantes :")

        for error in errors:
            print(f"  - {error}")

        print("\nGénération interrompue.")
        return 2

    if args.check:
        print("✓ Vérification terminée : aucune erreur bloquante.")
        return 0

    generate(
        rows,
        args.output,
        args.download,
        args.excel,
    )

    print(f"✓ Page générée : {args.output}")
    print(f"✓ Catalogue publié : {args.download}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""
V2 — Génère la page /methodes/ de CyberRepères à partir du catalogue
Excel « Méthodes, modèles et ressources méthodologiques utiles à la cybersécurité ».

Le script est volontairement spécialisé : il ne modifie pas
generate_referentiels.py.

Usage :
    python scripts/generate_methodes.py
    python scripts/generate_methodes.py --check
    python scripts/generate_methodes.py --excel chemin/vers/catalogue.xlsx
"""

from __future__ import annotations

import argparse
import html
import re
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_OUTPUT = Path("methodes/index.md")
DEFAULT_DOWNLOAD = Path("downloads/CyberReperes_Catalogue_Methodes.xlsx")

# Le nom exact du fichier peut évoluer : le script recherche automatiquement
# un classeur contenant « Outils » si le chemin par défaut n'existe pas.
DEFAULT_EXCEL = Path("catalogues/CyberReperes_Outils_et_Referentiels_interactifs.xlsx")

# Feuilles de travail : elles deviennent des catégories sur la page.
CATEGORY_SHEETS = [
    "Menaces",
    "Maturité",
    "Vulnérabilités",
    "Durcissement",
    "Évaluation",
    "Architecture",
    "Architecture logicielle",
    "Architecture entreprise",
    "Santé",
]

# Feuilles de documentation / correspondance, pas des catalogues.
IGNORED_SHEETS = {"Familles", "Lisez-moi"}


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def html_escape(value) -> str:
    return html.escape(clean(value), quote=True)


def is_url(value: str) -> bool:
    return bool(re.match(r"^https?://", clean(value), re.I))


def find_excel(explicit: Path | None) -> Path:
    if explicit:
        return explicit

    if DEFAULT_EXCEL.exists():
        return DEFAULT_EXCEL

    candidates = []
    for root in (Path("catalogues"), Path(".")):
        if not root.exists():
            continue
        for path in root.glob("*.xlsx"):
            name = path.name.lower()
            if "outil" in name or "interactif" in name:
                candidates.append(path)

    if len(candidates) == 1:
        return candidates[0]

    if not candidates:
        raise FileNotFoundError(
            "Catalogue Excel Outils introuvable. "
            f"Chemin attendu : {DEFAULT_EXCEL}"
        )

    raise FileNotFoundError(
        "Plusieurs catalogues Outils trouvés. Utilisez --excel pour préciser "
        "le fichier à utiliser :\n  " + "\n  ".join(map(str, candidates))
    )


def normalize_header(value: str) -> str:
    value = clean(value).lower()
    value = re.sub(r"[’']", "'", value)
    value = re.sub(r"\s+", " ", value)
    return value


def read_sheet(ws):
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return [], []

    headers = [clean(v) for v in values[0]]
    rows = []

    for raw in values[1:]:
        row = {headers[i]: clean(raw[i]) if i < len(raw) else ""
               for i in range(len(headers))}
        if any(row.values()):
            rows.append(row)

    return headers, rows


def find_column(headers, candidates):
    normalized = {normalize_header(h): h for h in headers}
    for candidate in candidates:
        c = normalize_header(candidate)
        if c in normalized:
            return normalized[c]
    return None


def pick(row, headers, candidates):
    col = find_column(headers, candidates)
    return clean(row.get(col, "")) if col else ""


def detail_fields(row, headers, excluded):
    """Retourne les colonnes non vides utiles dans le détail."""
    result = []
    for header in headers:
        if header in excluded:
            continue
        value = clean(row.get(header, ""))
        if value:
            result.append((header, value))
    return result


def make_link(label: str, value: str) -> str:
    value = clean(value)
    if not value:
        return ""
    if is_url(value):
        return f'<a href="{html_escape(value)}" target="_blank" rel="noopener">{html_escape(label)}</a>'
    return html_escape(value)


def render_category(sheet_name: str, headers, rows):
    title = html_escape(sheet_name)

    if not rows:
        return f"""
<section class="catalogue-section">
  <h2>{title}</h2>
  <p>Aucune entrée renseignée dans cet onglet.</p>
</section>
"""

    ref_col = find_column(headers, [
        "ID", "Référence", "Reference", "Réf.", "Ref", "Identifiant"
    ])
    name_col = find_column(headers, [
        "Outil / référentiel",
        "Modèle / outil",
        "Nom",
        "Outil",
        "Référentiel",
        "Titre",
        "Libellé",
    ])
    desc_col = find_column(headers, [
        "Objet",
        "Description",
        "Présentation",
        "Résumé",
        "Objectif",
        "Description courte",
    ])
    url_col = find_column(headers, [
        "Source officielle",
        "URL",
        "URL officielle",
        "Lien",
        "Site officiel",
        "Lien officiel",
    ])

    first_col = headers[0] if headers else None

    # Ces champs sont déjà visibles dans la ligne repliée.
    excluded = {c for c in [ref_col, name_col, desc_col] if c}

    out = [
        '<section class="catalogue-section">',
        f'  <h2>{title}</h2>',
        '  <div class="catalogue-table catalogue-tools-table">',
        '    <div class="catalogue-header">',
        '      <div>ID</div>',
        '      <div>Outil / référentiel</div>',
        '      <div>Objet</div>',
        '    </div>',
    ]

    for index, row in enumerate(rows, start=1):
        ref = clean(row.get(ref_col, "")) if ref_col else f"{index:03d}"
        name = clean(row.get(name_col, "")) if name_col else clean(row.get(first_col, ""))
        desc = clean(row.get(desc_col, "")) if desc_col else ""
        url = clean(row.get(url_col, "")) if url_col else ""

        # Le lien officiel est directement accessible depuis la ligne repliée.
        if url and is_url(url):
            name_html = (
                f'<a class="catalogue-tool-link" href="{html_escape(url)}" '
                f'target="_blank" rel="noopener" '
                f'aria-label="Ouvrir la source officielle de {html_escape(name)}">'
                f'{html_escape(name)} ↗</a>'
            )
        else:
            name_html = html_escape(name)

        details = detail_fields(row, headers, excluded)

        out.append('    <details class="catalogue-entry">')
        out.append('      <summary class="catalogue-row">')
        out.append(f'        <div class="catalogue-ref"><strong>{html_escape(ref)}</strong></div>')
        out.append(f'        <div>{name_html}</div>')
        out.append(f'        <div>{html_escape(desc)}</div>')
        out.append('      </summary>')

        out.append('      <div class="catalogue-detail">')

        for header, value in details:
            if is_url(value):
                rendered = make_link(value, value)
            else:
                rendered = html_escape(value).replace("\n", "<br>")
            out.append(
                f'        <p><strong>{html_escape(header)} :</strong> {rendered}</p>'
            )

        # La source officielle reste également disponible dans le détail.
        if url and url_col:
            rendered = make_link(url, url)
            out.append(
                f'        <p><strong>{html_escape(url_col)} :</strong> {rendered}</p>'
            )

        out.append('      </div>')
        out.append('    </details>')

    out.extend([
        '  </div>',
        '</section>',
        ''
    ])
    return "\n".join(out)


def load_catalogue(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)

    available = list(wb.sheetnames)
    selected = [s for s in CATEGORY_SHEETS if s in available]

    # Ne pas perdre silencieusement de nouveaux onglets : ils sont signalés.
    unknown = [
        s for s in available
        if s not in CATEGORY_SHEETS and s not in IGNORED_SHEETS
    ]

    categories = []
    for sheet in selected:
        headers, rows = read_sheet(wb[sheet])
        categories.append((sheet, headers, rows))

    return categories, available, unknown


def validate(categories, unknown):
    errors = []
    warnings = []

    if not categories:
        errors.append("Aucun onglet catalogue reconnu dans le fichier Excel.")

    for sheet, headers, rows in categories:
        if not headers:
            warnings.append(f"Onglet vide ou sans en-têtes : {sheet}")
            continue

        url_col = find_column(headers, [
            "URL", "URL officielle", "Lien", "Site officiel", "Lien officiel"
        ])

        if url_col:
            for row in rows:
                value = clean(row.get(url_col, ""))
                if value and not is_url(value):
                    ref = pick(row, headers, [
                        "Référence", "Reference", "Réf.", "Ref", "ID", "Nom"
                    ]) or "entrée sans référence"
                    warnings.append(
                        f"URL à vérifier : {sheet} / {ref} ({value})"
                    )

    for sheet in unknown:
        warnings.append(
            f"Onglet non traité : {sheet} "
            "(à classer dans CATEGORY_SHEETS ou IGNORED_SHEETS si nécessaire)."
        )

    return errors, warnings


def generate(categories, output: Path, download: Path, source_excel: Path):
    output.parent.mkdir(parents=True, exist_ok=True)
    download.parent.mkdir(parents=True, exist_ok=True)

    sections = []
    total = 0

    for sheet, headers, rows in categories:
        sections.append(render_category(sheet, headers, rows))
        total += len(rows)

    text = f"""---
layout: default
title: "Méthodes"
description: "Méthodes, modèles et ressources méthodologiques utiles à la cybersécurité."
nav: true
nav_order: 30
---

{{% include methodes-header.md %}}

**{total} entrées** · **{len(categories)} catégories**

📥 [Télécharger le catalogue Excel]({{{{ site.baseurl }}}}/{download.as_posix()})

{"".join(sections)}
"""

    output.write_text(text.rstrip() + "\n", encoding="utf-8")

    # Copie du classeur source dans le dossier public, sans modifier le maître.
    if source_excel.resolve() != download.resolve():
        download.write_bytes(source_excel.read_bytes())


def main():
    parser = argparse.ArgumentParser(
        description="Génère la page CyberRepères /methodes/ depuis le catalogue Excel."
    )
    parser.add_argument("--excel", type=Path, default=None,
                        help="Classeur Excel source.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT,
                        help="Page Markdown générée.")
    parser.add_argument("--download", type=Path, default=DEFAULT_DOWNLOAD,
                        help="Copie publique du catalogue Excel.")
    parser.add_argument("--check", action="store_true",
                        help="Vérifie le catalogue sans générer la page.")
    args = parser.parse_args()

    try:
        source_excel = find_excel(args.excel)
        categories, available, unknown = load_catalogue(source_excel)
    except Exception as exc:
        print(f"ERREUR : {exc}")
        return 2

    errors, warnings = validate(categories, unknown)

    print(f"Catalogue : {source_excel}")
    print(f"Onglets disponibles : {', '.join(available)}")
    print(f"Catégories traitées : {len(categories)}")

    for sheet, headers, rows in categories:
        print(f"  - {sheet} : {len(rows)} entrées")

    for warning in warnings:
        print(f"ATTENTION : {warning}")

    if errors:
        print("\nErreurs :")
        for error in errors:
            print(f"  - {error}")
        print("\nGénération interrompue.")
        return 2

    if args.check:
        print("✓ Vérification terminée : aucune erreur bloquante.")
        return 0

    generate(categories, args.output, args.download, source_excel)
    print(f"✓ Page générée : {args.output}")
    print(f"✓ Catalogue publié : {args.download}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
